import pandas as pd
import os
from openpyxl import load_workbook

def convert_time_to_score_and_comment(file_path, max_score=100, min_score=90, threshold_days=5):
    """
    从Excel中提取三列，将领取时间转为相对秒数并计算分数，生成作业评语，返回DataFrame
    """
    df = pd.read_excel(file_path, header=1)
    desired_columns = ['学号/工号', '学生姓名', '领取时间']
    df_selected = df[desired_columns].copy()

    # 转换时间戳
    df_selected['领取时间'] = pd.to_datetime(df_selected['领取时间'], errors='coerce')

    # 计算相对秒数
    min_time = df_selected['领取时间'].min()
    df_selected['相对领取时间_秒'] = df_selected['领取时间'].apply(
        lambda x: float('inf') if pd.isna(x) else (x - min_time).total_seconds()
    )

    # 计算最大秒数（非inf）
    finite_seconds = df_selected.loc[df_selected['相对领取时间_秒'] != float('inf'), '相对领取时间_秒']
    max_seconds = finite_seconds.max()

    # 分数列
    def calculate_score(seconds):
        if seconds == 0:
            return max_score
        elif seconds == float('inf'):
            return 0
        else:
            score = min_score + (max_score - min_score) * (1 - seconds / max_seconds)
            return int(score)

    df_selected['分数'] = df_selected['相对领取时间_秒'].apply(calculate_score)

    # 作业评语逻辑
    median_seconds = finite_seconds.median()
    threshold_seconds = max(median_seconds, threshold_days * 24 * 3600)

    def generate_comment(seconds):
        if seconds == float('inf'):
            return '作业缺交'
        elif seconds == 0 or seconds < threshold_seconds:
            return '本次作业整体不错，继续努力'
        else:
            return '作业要抓紧时间完成'

    df_selected['作业批语'] = df_selected['相对领取时间_秒'].apply(generate_comment)

    return df_selected

def xls_to_xlsx(xls_path, xlsx_path):
    """
    将xls文件转为xlsx文件
    """
    data = pd.read_excel(xls_path, header=None)
    data.to_excel(xlsx_path, index=False, header=False)
    print(f'已将 {xls_path} 转换为 {xlsx_path}')

def update_xlsx_with_scores_and_comments(xlsx_path, df_final):
    """
    根据df_final更新xlsx文件中的【分数】【作业批语】
    """
    wb = load_workbook(xlsx_path)
    ws = wb.active

    update_map = {
        (str(row['学号/工号']).strip(), str(row['学生姓名']).strip()): (row['分数'], row['作业批语'])
        for _, row in df_final.iterrows()
    }

    for row in ws.iter_rows(min_row=2, values_only=False):
        id_cell = row[0]  # 学号/工号
        name_cell = row[1]  # 学生姓名

        id_val = str(id_cell.value).strip() if id_cell.value else ''
        name_val = str(name_cell.value).strip() if name_cell.value else ''

        key = (id_val, name_val)
        if key in update_map:
            score, comment = update_map[key]
            row[8].value = score  # 第9列：分数
            row[9].value = comment  # 第10列：作业批语

    wb.save(xlsx_path)
    print(f'已更新 {xlsx_path} 文件中的【分数】【作业批语】')

import pyexcel

def xlsx_to_xls_via_pyexcel(xlsx_path, xls_path):
    """
    使用 pyexcel 将 xlsx 文件转换为 xls 文件
    """
    pyexcel.save_book_as(file_name=xlsx_path, dest_file_name=xls_path)
    print(f'已将 {xlsx_path} 转换为 {xls_path}')


# 使用示例
xls_path = r'C:\Users\xijia\Downloads\课堂练习4.xls'
xlsx_path = r'C:\Users\xijia\Downloads\t上机实验.xlsx'

# 生成df_final
df_final = convert_time_to_score_and_comment(xls_path, max_score=100, min_score=85, threshold_days=5)

# 转为xlsx
xls_to_xlsx(xls_path, xlsx_path)
os.remove(xls_path)

# 更新xlsx中的分数和作业批语
update_xlsx_with_scores_and_comments(xlsx_path, df_final)

# 转回xls
xlsx_to_xls_via_pyexcel(xlsx_path, xls_path)
os.remove(xlsx_path)

